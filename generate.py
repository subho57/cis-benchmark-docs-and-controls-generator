import argparse
import os
import re
import sys
from pathlib import Path
from typing import Dict, Optional, Tuple

import jsonc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="CIS Benchmark Docs and Controls Generator for Steampipe",
        epilog="Example: cis-benchmark-generator --benchmark path/to/benchmark.xlsx --docs --controls --output output_dir"
    )
    parser.add_argument("--benchmark", required=True, help="Path to the CIS benchmark XLSX file")
    parser.add_argument("--docs", action="store_true", help="Generate documentation")
    parser.add_argument("--controls", action="store_true", help="Generate controls")
    parser.add_argument("--output", default="./", help="Output directory (default: ./)")
    return parser.parse_args()


def validate_benchmark_file(file_path):
    if not os.path.exists(file_path):
        print(f"Error: Benchmark file '{file_path}' does not exist.")
        sys.exit(1)
    if not file_path.lower().endswith('.xlsx'):
        print(f"Error: Benchmark file '{file_path}' is not an XLSX file.")
        sys.exit(1)


def extract_benchmark_info(filename: str) -> Tuple[Optional[str], Optional[str]]:
    # Pattern for filenames with underscores
    pattern1 = r'CIS_(.+)_Benchmark_v(\d+\.\d+\.\d+)(?:\s+\w+)?'

    # Pattern for filenames with spaces
    pattern2 = r'CIS (.+) Benchmark v(\d+\.\d+\.\d+)(?:\s+\w+)?'

    # Try the first pattern
    match = re.search(pattern1, filename)
    if match:
        return match.group(1).replace('_', ' '), match.group(2)

    # If first pattern doesn't match, try the second
    match = re.search(pattern2, filename)
    if match:
        return match.group(1), match.group(2)

    # If no pattern matches, return None
    return None, None


def parse_benchmark(file_path: str) -> Tuple[ReadOnlyWorksheet, ReadOnlyWorksheet, Optional[str], Optional[str]]:
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        # Add your parsing logic here
        print(f"Successfully loaded benchmark file: {file_path}")
        print(f"Sheets in the workbook: {workbook.sheetnames}")
        if 'Combined Profiles' not in workbook.sheetnames:
            print("Error: Combined Profiles sheet not found in the workbook.")
            sys.exit(1)
        benchmark_name, benchmark_version = extract_benchmark_info(os.path.basename(file_path))
        overview_worksheet = workbook['Overview - Glossary'] if 'Overview - Glossary' in workbook.sheetnames else None
        return workbook['Combined Profiles'], overview_worksheet, benchmark_name, benchmark_version
        # Return parsed data or perform further operations
    except Exception as e:
        print(f"Error parsing benchmark file: {e}")
        sys.exit(1)


def format_json_string(json_string: str) -> str:
    try:
        json_data = jsonc.loads(json_string)
        formatted_json_string = jsonc.dumps(json_data, indent=2)
        return formatted_json_string
    except jsonc.JSONDecodeError:
        print("Error: Invalid JSON string.", json_string)
        return json_string


def add_lang_to_code_blocks(markdown_string: str) -> str:
    parts = markdown_string.split("```")
    markdown_string = ""
    code_block_lang = "bash"
    for index, part in enumerate(parts):
        if index == len(parts) - 1:
            markdown_string += part
        elif index % 2 == 0:
            next_part = parts[index + 1]
            if next_part.startswith("\n{"):
                code_block_lang = "json"
            else:
                code_block_lang = "bash"
            markdown_string += f"{part}```{code_block_lang}" if part.endswith("\n\n") else f"{part}\n```{code_block_lang}"
        else:
            markdown_string += f"{part}```" if code_block_lang == "bash" else f"\n{format_json_string(part)}\n```"
            next_part = parts[index + 1]
            if not next_part.startswith("\n\n"):
                markdown_string += "\n"
    return markdown_string


def replace_remediation_headers(text: str) -> str:
    # Define the pattern to search for and its replacement
    patterns = [
        r'\*\*Remediate from (.*?)\:\*\*',
        r'\*\*Remediate from (.*?)\*\*',
        r'\*\*Remediation from (.*?)\:\*\*',
        r'\*\*Remediation from (.*?)\*\*',
        r'\*\*From (.*?)\:\*\*',
        r'\*\*From (.*?)\*\*',
        r'\*\*Audit from (.*?)\*\*',
        r'\*\*Audit from (.*?)\:\*\*',
    ]
    replacement = r'### From \1'

    for pattern in patterns:
        # Use re.sub to replace all occurrences
        text = re.sub(pattern, replacement, text)

    text = re.sub(r'\*\*Step (.*?)\*\*', r'### Step \1', text)

    return text


def is_line_item(line: str, ignore_unordered: bool = False) -> bool:
    striped_line = line.strip()
    first_char = striped_line[0] if len(striped_line) else ''
    second_char = striped_line[1] if len(striped_line) > 1 else ''
    third_char = striped_line[2] if len(striped_line) > 2 else ''
    is_item = first_char.isdigit() and second_char in [')', '.', ':']
    if not ignore_unordered and first_char in ['-', '*', '+'] and second_char == ' ':
        is_item = True
    if second_char.isdigit() and first_char.isdigit() and third_char in [')', '.', ':']:
        is_item = True
    return is_item


def renumber_markdown_lists(markdown_text):
    def replacement(match):
        nonlocal counter
        counter += 1
        replaced_str = f"{counter}.{match.group(1)}"
        if not replaced_str.endswith('.') and not replaced_str.endswith('.`') and not replaced_str.endswith(':') and not replaced_str.endswith(','):
            replaced_str += '.'
        return replaced_str

    lines = markdown_text.split('\n')
    numbered_lines = []
    counter = 0
    in_list = False

    for index, line in enumerate(lines):
        if re.match(r'1\.\s', line):
            if not in_list:
                counter = 0
                in_list = True
                previous_line = lines[index - 1]
                if previous_line != '':
                    numbered_lines.append('')
            numbered_line = re.sub(r'1\.(\s.*)', replacement, line)
            numbered_lines.append(numbered_line)
        else:
            if in_list and (line.strip() == '' or is_line_item(line, True)):
                in_list = False
            numbered_lines.append(line)

    return '\n'.join(numbered_lines)


def is_url(word: str) -> bool:
    return word.startswith('http://') or word.startswith('https://')


def convert_to_md_url(word: str) -> str:
    if is_url(word):
        word = word.removesuffix('.')
        return f"[{word}]({word})"
    return word


def trim_each_line(text: str) -> str:
    is_inside_code_block = False

    def trim_line(line: str) -> str:
        line = ' '.join([convert_to_md_url(word) for word in line.split(' ')])
        striped_line = line.strip()
        # if striped_line.startswith('**') and striped_line.endswith('**'):
        #     return f"### {striped_line[2:-2]}"
        nonlocal is_inside_code_block
        if line.strip().startswith('```'):
            is_inside_code_block = not is_inside_code_block
        if is_inside_code_block or line.strip().startswith('```'):
            return line.strip()
        if is_line_item(line, True) and not striped_line.endswith('.') and not striped_line.endswith(
                '.`') and not striped_line.endswith(':') and not striped_line.endswith(','):
            return line.rstrip() + '.'  # Add a period at the end of the line
        return line.rstrip()
    return '\n'.join([trim_line(line) for line in text.split('\n')])


def parse_each_benchmark(benchmark_data: pd.Series, benchmark_name: str, benchmark_version: str, generate_doc: bool = True,
                         generate_control: bool = True) -> Tuple[Optional[Dict[str, str]], Optional[str]]:
    if not (generate_doc or generate_control):
        return None, None
    section: str = benchmark_data['Section #']
    recommendation: str = benchmark_data['Recommendation #']
    profile: str = benchmark_data['Profile']
    title: str = benchmark_data['Title']
    status: str = benchmark_data['Assessment Status']
    description: str = benchmark_data['Description']
    rationale: str = benchmark_data['Rationale Statement']
    remidiation: str = benchmark_data['Remediation Procedure']
    default_value: str = benchmark_data['Default Value']

    if not (section and title):
        print(f"Skipping: {section} - {title}")
        return None, None

    if remidiation:
        remidiation = replace_remediation_headers(remidiation)

    if default_value and not default_value.endswith('.') and not default_value.endswith('.`') and len(default_value.split(' ')) != 1:
        default_value += '.'  # Add a period at the end if it's missing

    file_suffix = str(recommendation).replace(".", "_") if recommendation else str(section).replace(".", "_")
    doc_filename = f"cis_v{benchmark_version.replace('.', '')}_{file_suffix}.md"

    print(f"Processing: {recommendation or section} - {title}")
    markdown: Optional[str] = None

    if generate_doc:
        markdown = "## Description" if remidiation else "## Overview"
        if description:
            markdown += f"\n\n{description}"
        else:
            markdown += f"\n\nThis section covers security recommendations for {title}."
        if rationale:
            markdown += f"\n\n{rationale}"
        if remidiation:
            markdown += f"\n\n## Remediation\n\n{remidiation}"
        if default_value:
            markdown += f"\n\n### Default Value\n\n{default_value}"
        if markdown:
            markdown = trim_each_line(markdown)
            markdown = add_lang_to_code_blocks(markdown)
            markdown = renumber_markdown_lists(markdown)
            if not markdown.endswith('\n'):
                markdown += "\n"

    control: Optional[str] = None
    if generate_control:
        pass
    doc = {
        "filename": doc_filename,
        "markdown": markdown
    }
    return doc, control


def generate_overview_doc(overview_data: ReadOnlyWorksheet, benchmark_name: str) -> str:
    levels = {}
    overview = None
    rows = list(overview_data.iter_rows(values_only=True))
    for index, row in enumerate(rows):
        value: str = row[0]
        if value and value.startswith('Level '):
            levels[value] = rows[index + 1][0]
        if value == 'Overview':
            overview = rows[index + 1][0].splitlines()[0]
    if overview:
        overview = overview.replace('This spreadsheet', f'The CIS {benchmark_name} Benchmark')
    markdown = f"To obtain the latest version of the official guide, please visit http://benchmarks.cisecurity.org.\n\n## Overview\n\n{overview}\n\n## Profiles\n\n"
    level_descriptions = "\n\n".join([f"### {level}\n\n{description}" for level, description in levels.items()])
    markdown += level_descriptions.replace('  ', ' ')
    if not markdown.endswith('\n'):
        markdown += "\n"
    markdown = trim_each_line(markdown)
    markdown = add_lang_to_code_blocks(markdown)
    print(f"Generating Overview documentation for {benchmark_name}")
    return markdown


def generate_docs_and_controls(
        profiles_worksheet: ReadOnlyWorksheet,
        overview_data: Optional[ReadOnlyWorksheet],
        output_dir: str,
        benchmark_name: str,
        benchmark_version: str,
        generate_docs: bool = True,
        generate_controls: bool = True):
    data = list(profiles_worksheet.values)
    headers = data[0]
    data = data[1:]
    # Create a DataFrame
    df = pd.DataFrame(data, columns=headers)

    output_path = Path(output_dir) / f"cis_v{benchmark_version.replace('.', '')}" / "docs"
    output_path.mkdir(parents=True, exist_ok=True)

    if generate_docs and overview_data:
        overview = generate_overview_doc(overview_data, benchmark_name)
        Path(output_path / "cis_overview.md").write_text(overview)

    benchmarks = [parse_each_benchmark(row, benchmark_name, benchmark_version, generate_docs, generate_controls) for _, row in df.iterrows()]
    for i, (doc, control) in enumerate(benchmarks):
        if doc:
            filepath = output_path / doc["filename"]
            filepath.write_text(doc["markdown"])
        if control:
            with open(os.path.join(output_dir, f"control_{i}.sql"), "w") as f:
                f.write(control)


def main():
    args = parse_arguments()

    if not (args.docs or args.controls):
        print("Warning: Neither --docs nor --controls specified. No output generated.")
        sys.exit(0)

    validate_benchmark_file(args.benchmark)
    parsed_data, overview, benchmark_name, benchmark_version = parse_benchmark(args.benchmark)

    if not benchmark_name or not benchmark_version:
        print("Error: Could not extract benchmark name and version from the filename. Ensure the filename follows the CIS Benchmark naming convention.")
        sys.exit(1)

    os.makedirs(args.output, exist_ok=True)

    generate_docs_and_controls(parsed_data, overview, args.output, benchmark_name, benchmark_version, args.docs, args.controls)

    print(f"Output generated in directory: {args.output}")


if __name__ == "__main__":
    main()
